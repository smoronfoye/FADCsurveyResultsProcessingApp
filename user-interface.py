from PyQt5.QtWidgets import (QApplication, QFormLayout, QGroupBox,
                             QLabel, QPushButton, QVBoxLayout, QWidget,
                             QMainWindow, QLineEdit, QFileDialog, QComboBox)
from PyQt5.QtCore import Qt
from processSurveyResultsFile import processExcelFile
from openpyxl import load_workbook
import pandas as pd
import re

class MainWindow(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.createUI()


    def createUI(self):
        # Create window
        self.setWindowTitle('FADC Survey Results File Processor')
        self.resize(700, 400)
        self.setMinimumSize(250, 300)
        # Create central widget and layout
        self._centralWidget = QWidget()
        self._verticalLayout = QVBoxLayout()
        self._centralWidget.setLayout(self._verticalLayout)
        # Set central widget
        self.setCentralWidget(self._centralWidget)
        # Vertically center widgets
        self._verticalLayout.addStretch(1)
        self.addText()
        self.addInputText()
        # Vertically center widgets
        self._verticalLayout.addStretch(1)



    def addText(self):
        messageLabel = QLabel(
            'Please enter the name of the Excel file and the Excel Sheet in the fields below'
        )
        messageLabel.setAlignment(Qt.AlignCenter)
        messageLabel.setFixedWidth(350)
        messageLabel.setMinimumHeight(50)
        messageLabel.setWordWrap(True)
        self._verticalLayout.addWidget(messageLabel, alignment=Qt.AlignCenter)


    def addInputText(self):
        self.groupBox = QGroupBox()
        self.groupBox.setFixedWidth(550)

        formLayout = QFormLayout()

        self.excelFileNameLabel = QLabel('Excel File Name')
        self.selectFileButton = QPushButton('Click To Select File', self)
        self.selectFileButton.setMaximumWidth(150)
        self.selectFileButton.clicked.connect(lambda: self.openFileNameDialog())

        self.excelFileNameField = QLineEdit()
        self.excelFileNameField.setTextMargins(3, 0, 3, 0)
        self.excelFileNameField.setMinimumWidth(200)
        self.excelFileNameField.setMaximumWidth(300)
        self.excelFileNameField.setClearButtonEnabled(True)
        self.excelFileNameField.setEnabled(False)

        self.excelSheetNameLabel = QLabel('Excel Sheet Name')
        self.selectSheetName = QComboBox()
        self.selectSheetName.setContentsMargins(3, 0, 3, 0)
        self.selectSheetName.setMinimumWidth(200)
        self.selectSheetName.setMaximumWidth(300)
        self.selectSheetName.currentTextChanged.connect(self.onSelectedSheetnameChanged)

        self.latitudeColumnLabel = QLabel('Latitude Column')
        self.selectLatitudeColumn = QComboBox()
        self.selectLatitudeColumn.setContentsMargins(3, 0, 3, 0)
        self.selectLatitudeColumn.setMinimumWidth(200)
        self.selectLatitudeColumn.setMaximumWidth(300)

        self.longitudeColumnLabel = QLabel('Longitude Column')
        self.selectLongitudeColumn = QComboBox()
        self.selectLongitudeColumn.setContentsMargins(3, 0, 3, 0)
        self.selectLongitudeColumn.setMinimumWidth(200)
        self.selectLongitudeColumn.setMaximumWidth(300)

        self.processButton = QPushButton('Process File', self)
        self.processButton.setMaximumWidth(100)
        self.processButton.clicked.connect(lambda: self.process_button_click())

        self.statusLabel = QLabel('')

        formLayout.addRow(self.selectFileButton, self.excelFileNameField)
        formLayout.addRow(self.excelSheetNameLabel, self.selectSheetName)
        formLayout.addRow(self.latitudeColumnLabel, self.selectLatitudeColumn)
        formLayout.addRow(self.longitudeColumnLabel, self.selectLongitudeColumn)
        formLayout.addRow(self.processButton)
        formLayout.addRow(self.statusLabel)

        self.groupBox.setLayout(formLayout)
        self._verticalLayout.addWidget(self.groupBox, alignment=Qt.AlignCenter)

    def process_button_click(self):
        
        fileName = self.excelFileNameField.text()
        sheetName = self.selectSheetName.currentText()
        latitudeColumn = self.selectLatitudeColumn.currentText()
        longitudeColumn = self.selectLongitudeColumn.currentText()
        if fileName is not None and fileName != '' and sheetName is not None and sheetName != '' and latitudeColumn is not None and latitudeColumn != '' and longitudeColumn is not None and longitudeColumn != '': 
            self.statusLabel.setText('Loading Please Wait...')
            self.statusLabel.repaint()
            try:
                processExcelFile(fileName, sheetName, latitudeColumn, longitudeColumn)
            except:
                self.statusLabel.setText('Error Encountered while processing. Please check all inputs')
                self.statusLabel.repaint()
            self.statusLabel.setText('Processing Completed!')
            self.statusLabel.repaint()
        else:
            self.statusLabel.setText('Error Encountered while processing. Please make sure all inputs are provided')
            self.statusLabel.repaint()

    def openFileNameDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","All Files (*);;Python Files (*.py)", options=options)
        if fileName:
            self.excelFileNameField.setText(fileName)
            wb = load_workbook(fileName, read_only=True, keep_links=False)
            self.selectSheetName.addItems(wb.sheetnames)

    def onSelectedSheetnameChanged(self, sheetName):
        self.selectLatitudeColumn.clear()
        self.selectLongitudeColumn.clear()
        fileName = self.excelFileNameField.text()

        df = pd.read_excel(fileName, sheet_name=sheetName)
        dataColumnNames = df.columns.tolist()
        dataColumnNames.insert(0, 'Please Select And Item')
        self.selectLatitudeColumn.addItems(dataColumnNames)
        self.selectLongitudeColumn.addItems(dataColumnNames)

        for columnName in dataColumnNames:
            
            if re.search('latitude', columnName, re.IGNORECASE):
                index = self.selectLatitudeColumn.findText(columnName, Qt.MatchFixedString)
                if index >= 0:
                    self.selectLatitudeColumn.setCurrentIndex(index)
            
            elif re.search('longitude', columnName, re.IGNORECASE):
                index = self.selectLongitudeColumn.findText(columnName, Qt.MatchFixedString)
                if index >= 0:
                    self.selectLongitudeColumn.setCurrentIndex(index)
        



if (__name__ == '__main__'):

    application = QApplication([])
    mainWindow = MainWindow()
    mainWindow.show()
    application.exec()